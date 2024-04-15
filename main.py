import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import load_workbook
from tabulate import tabulate

def view():
    workbook = load_workbook(filename='WhiteCardPatients.xlsx')
    sheet = workbook.active

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if all(cell is not None for cell in row):
            data.append((row[0], row[1], row[2],row[-2], row[-1]))  # Select first three and last columns

    workbook.close()

    output_text.delete(1.0, tk.END)  # Clear previous output
    output_text.insert(tk.END, tabulate(data, headers=["Surname", "First Initial", "Address", "Telephone" , "Reference Number"], tablefmt="pretty"))

def sort_workbook():
    workbook = load_workbook(filename='WhiteCardPatients.xlsx')
    sheet = workbook.active

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None :
            data.append((row[0].strip(), row[1], row[2], row[-2],row[-1]))  # Select first three and last two columns

    data.sort(key=lambda x: x[0])
    workbook.close()

    output_text.delete(1.0, tk.END)  # Clear previous output
    output_text.insert(tk.END, "Sorted records:\n")
    output_text.insert(tk.END, tabulate(data, headers=["Surname", "First Initial", "Address","Telephone" ,"Reference Number"], tablefmt="pretty"))


def search_popup():
    search_window = tk.Toplevel(root)
    search_window.title("Search Record")
    search_window.geometry("400x300")  # Set the size of the search window

    # Create labels and entry widgets for search criteria
    surname_label = ttk.Label(search_window, text="Surname:")
    surname_label.grid(row=0, column=0, padx=5, pady=5)
    surname_entry = ttk.Entry(search_window)
    surname_entry.grid(row=0, column=1, padx=5, pady=5)

    fInitial_label = ttk.Label(search_window, text="First Initial:")
    fInitial_label.grid(row=1, column=0, padx=5, pady=5)
    fInitial_entry = ttk.Entry(search_window)
    fInitial_entry.grid(row=1, column=1, padx=5, pady=5)

    address_label = ttk.Label(search_window, text="Address:")
    address_label.grid(row=2, column=0, padx=5, pady=5)
    address_entry = ttk.Entry(search_window)
    address_entry.grid(row=2, column=1, padx=5, pady=5)

    reference_label = ttk.Label(search_window, text="Reference Number:")
    reference_label.grid(row=3, column=0, padx=5, pady=5)
    reference_entry = ttk.Entry(search_window)
    reference_entry.grid(row=3, column=1, padx=5, pady=5)

    # Create the Search button
    search_button = ttk.Button(search_window, text="Search Record",
                               command=lambda: search(surname_entry.get(), fInitial_entry.get(), address_entry.get(),
                                                      reference_entry.get()))
    search_button.grid(row=4, column=0, columnspan=2, pady=10)

    # Create the Delete and Modify buttons
    delete_button = ttk.Button(search_window, text="Delete Record", command=lambda: delete(reference_entry.get()))
    delete_button.grid(row=5, column=0, padx=5, pady=5)

    modify_button = ttk.Button(search_window, text="Modify Record", command=lambda: modify_popup(reference_entry.get()))
    modify_button.grid(row=5, column=1, padx=5, pady=5)

    # Center the search window on the screen
    search_window.update_idletasks()
    width = search_window.winfo_width()
    height = search_window.winfo_height()
    x_offset = (search_window.winfo_screenwidth() - width) // 2
    y_offset = (search_window.winfo_screenheight() - height) // 2
    search_window.geometry(f"+{x_offset}+{y_offset}")

    # Move the reference number entry and buttons below the search button
    reference_label.grid(row=6, column=0, padx=5, pady=5)
    reference_entry.grid(row=6, column=1, padx=5, pady=5)
    delete_button.grid(row=7, column=0, padx=5, pady=5)
    modify_button.grid(row=7, column=1, padx=5, pady=5)

def search(searchSurname, searchFInitial, searchAddress, searchReference):

    # Remove leading and trailing spaces and convert to lowercase if searchSurname is not None
    searchSurname = searchSurname.strip().lower() if searchSurname else ""
    # Remove leading and trailing spaces and convert to lowercase if searchFInitial is not None
    searchFInitial = searchFInitial.strip().lower() if searchFInitial else ""
    # Remove leading and trailing spaces and convert to lowercase if searchAddress is not None
    searchAddress = searchAddress.strip().lower() if searchAddress else ""
    # Remove leading and trailing spaces and convert to lowercase if searchReference is not None
    searchReference = searchReference.strip().lower() if searchReference else ""

    #if all of the textboxes are empty, then display error
    if searchSurname.strip() == "" and searchFInitial.strip() == "" and searchAddress == "":
        output_text.delete(1.0, tk.END)  # Clear previous output
        messagebox.showerror(tk.END, "Please fill at least one textbox\n")
        return

    workbook_filename = 'WhiteCardPatients.xlsx'  # Filename
    workbook = load_workbook(filename=workbook_filename)
    sheet = workbook.active

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if ((row[0] is not None and row[1] is not None and row[2] is not None) and
                (str(row[0]).lower() == searchSurname and str(row[1]).lower() == searchFInitial and str(row[2]).lower() == searchAddress) or
                (searchSurname in str(row[0]).lower() and searchFInitial in str(row[1]).lower() and searchAddress in str(row[2]).lower())):

            # Append only the first 3 and last 2 columns to the data list
            data.append(row[:3] + row[-2:])  # Select first three and last two columns

    workbook.close()

    output_text.delete(1.0, tk.END)  # Clear previous output
    if data:
        headers = ["Surname", "First Initial", "Address", "Telephone", "Reference No."]
        output_text.insert(tk.END, tabulate(data, headers=headers, tablefmt="pretty"))
    else:
        output_text.insert(tk.END, "Record doesn't exist\n")

def insert_popup():
    # Create the insert window
    insert_window = tk.Toplevel(root)
    insert_window.title("Insert Record")
    insert_window.geometry("400x300")  # Set the size of the insert window

    # Calculate the center of the window
    window_width = insert_window.winfo_reqwidth()
    window_height = insert_window.winfo_reqheight()
    position_right = int(insert_window.winfo_screenwidth() / 2 - window_width / 2)
    position_down = int(insert_window.winfo_screenheight() / 2 - window_height / 2)

    # Set the window position
    insert_window.geometry("+{}+{}".format(position_right, position_down))

    # Create labels and entry widgets for inserting a record
    surname_label = ttk.Label(insert_window, text="Surname:")
    surname_label.grid(row=0, column=0, padx=5, pady=5)
    surname_entry = ttk.Entry(insert_window)
    surname_entry.grid(row=0, column=1, padx=5, pady=5)

    fInitial_label = ttk.Label(insert_window, text="First Initial:")
    fInitial_label.grid(row=1, column=0, padx=5, pady=5)
    fInitial_entry = ttk.Entry(insert_window)
    fInitial_entry.grid(row=1, column=1, padx=5, pady=5)

    address_label = ttk.Label(insert_window, text="Address:")
    address_label.grid(row=2, column=0, padx=5, pady=5)
    address_entry = ttk.Entry(insert_window)
    address_entry.grid(row=2, column=1, padx=5, pady=5)

    telephone_label = ttk.Label(insert_window, text="Telephone:")
    telephone_label.grid(row=3, column=0, padx=5, pady=5)
    telephone_entry = ttk.Entry(insert_window)
    telephone_entry.grid(row=3, column=1, padx=5, pady=5)

    reference_label = ttk.Label(insert_window, text="Reference No.:")
    reference_label.grid(row=4, column=0, padx=5, pady=5)
    reference_entry = ttk.Entry(insert_window)
    reference_entry.grid(row=4, column=1, padx=5, pady=5)

    # Add save record button
    save_button = ttk.Button(insert_window, text="Save Record",
                             command=lambda: save_record(surname_entry, fInitial_entry, address_entry, telephone_entry,reference_entry,
                                                         insert_window))
    save_button.grid(row=5, column=0, columnspan=2)

def save_record(surname_entry, fInitial_entry, address_entry, telephone_entry, reference_entry, insert_window):
    # Load the workbook
    workbook_filename = 'WhiteCardPatients.xlsx'  # Filename
    workbook = load_workbook(filename=workbook_filename)
    sheet = workbook.active

    # Calculate reference number
    surname = surname_entry.get()
    f_initial = fInitial_entry.get()
    address = address_entry.get()
    telephone = telephone_entry.get()
    reference_number = reference_entry.get()

    found = "Maybe"

    if (surname =="" or f_initial =="" or address =="" or telephone =="" or reference_number == "") or (surname =="" and f_initial =="" and address =="" and telephone =="" and reference_number == "")  :
        messagebox.showerror("Error","Please fill in all textboxes!")
        insert_window.destroy()
          # USed to determine whether an existing reference number matches, the refernce number entered
    else:

        # Iterate over the whole to check if there are existing records was the inserted reference number
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if str(row[7]).strip() == str(reference_number).strip():  # Assuming reference number is in the 8th column (index 7)
                found = "Yes"
                break # exit loop once an existing reference number is found
            else:
                found = "No"
        #End of for-loop

        if found =="Yes":
            messagebox.showerror("Error", "Reference number already exists!")
            # Close the workbook
            workbook.close()

        elif found =="No":
            # Create a new row with the data if entered reference number doesn't exist already
            new_row = [
                surname,
                f_initial,
                address,
                "",
                "",
                "",
                telephone,
                reference_number
            ]

            # Append the new row to the next available row after the last existing row
            sheet.append(new_row)

            # Save the workbook
            workbook.save(filename=workbook_filename)

            # Close the workbook
            workbook.close()
            # Display success message and close insert window
            messagebox.showinfo("Record inserted", "Record inserted successfully")
            insert_window.destroy()

def delete(reference_number):
    workbook = load_workbook(filename='WhiteCardPatients.xlsx')
    sheet = workbook.active

    found = "No"  # Flag to track if the record was found and deleted

    # Iterate over rows and find the row with the matching reference number, then delete it
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row[7] == reference_number:  # Assuming reference number is in the 8th column (index 7)
            result = messagebox.askquestion("Delete", f"Are you sure you want to delete {reference_number}?")

            if result =="yes":

                sheet.delete_rows(row_index)  # Delete the entire row using row index directly
                found = "Yes"

            elif result == "no":
                found = "Maybe"
                break  # Exit loop after deleting the first occurrence




    workbook.save(filename='WhiteCardPatients.xlsx')  # Save the workbook after deletion
    workbook.close()  # Close the workbook

    if found == "Yes":
        messagebox.showinfo("Record Deleted", f"Record with reference number {reference_number} deleted successfully")
    elif found == "No":
        messagebox.showerror("Record Not Found", f"Record with reference number {reference_number} not found")
    elif found == "Maybe":
        messagebox.showinfo("", f"Delete Cancelled for {reference_number}")

def modify_popup(reference_number):
    modify_window = tk.Toplevel(root)
    modify_window.title("Modify Record")
    modify_window.geometry("400x300")  # Set the size of the modify window

    # Add entry widgets for modifying record
    surname_label = ttk.Label(modify_window, text="Surname:")
    surname_label.grid(row=0, column=0,sticky='nsew')
    surname_entry = ttk.Entry(modify_window)
    surname_entry.grid(row=0, column=1,sticky='nsew')

    fInitial_label = ttk.Label(modify_window, text="First Initial:")
    fInitial_label.grid(row=1, column=0,sticky='nsew')
    fInitial_entry = ttk.Entry(modify_window)
    fInitial_entry.grid(row=1, column=1,sticky='nsew')

    address_label = ttk.Label(modify_window, text="Address:")
    address_label.grid(row=2, column=0,sticky='nsew')
    address_entry = ttk.Entry(modify_window)
    address_entry.grid(row=2, column=1,sticky='nsew')

    telephone_label = ttk.Label(modify_window, text="Telephone:")
    telephone_label.grid(row=6, column=0,sticky='nsew')
    telephone_entry = ttk.Entry(modify_window)
    telephone_entry.grid(row=6, column=1,sticky='nsew')

    reference_number_label = ttk.Label(modify_window, text="Reference Number:")
    reference_number_label.grid(row=7, column=0,sticky='nsew')
    reference_number_entry = ttk.Entry(modify_window)
    reference_number_entry.grid(row=7, column=1,sticky='nsew')
    reference_number_entry.insert(0, reference_number)  # Insert the reference number passed as an argument

    # Function to update record
    def update_record():
        surname = surname_entry.get()
        f_initial = fInitial_entry.get()
        address = address_entry.get()
        town = ""
        city = ""
        area_code = ""
        telephone = telephone_entry.get()

        # Load the workbook
        workbook = load_workbook(filename='WhiteCardPatients.xlsx')
        sheet = workbook.active

        found = "No"  # Flag to track if the record was found and updated

        # Iterate over rows and find the row with the matching reference number
        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if row[7] == reference_number:  # Assuming reference number is in the 8th column (index 7)

                result= messagebox.askquestion("Modify Record",f"Do you want to modify {reference_number} ?")
                if result =="yes":
                    # Update the record with the new values if the corresponding entry is not blank
                    if surname:
                        sheet.cell(row=row_index, column=1).value = surname
                    if f_initial:
                        sheet.cell(row=row_index, column=2).value = f_initial
                    if address:
                        sheet.cell(row=row_index, column=3).value = address
                    if town:
                        sheet.cell(row=row_index, column=4).value = town
                    if city:
                        sheet.cell(row=row_index, column=5).value = city
                    if area_code:
                        sheet.cell(row=row_index, column=6).value = area_code
                    if telephone:
                        sheet.cell(row=row_index, column=7).value = telephone
                    found = "Yes"
                elif result =="no":
                    found = "Maybe"
                    break  # Exit loop after updating the first occurrence

        if found == "Yes":
            workbook.save(filename='WhiteCardPatients.xlsx')  # Save the workbook after updating
            workbook.close()  # Close the workbook
            messagebox.showinfo("Record Updated",
                                f"{reference_number} updated successfully")
            modify_window.destroy()  # Close the modify window

        elif found== "No":
            workbook.close()  # Close the workbook
            messagebox.showerror("Record Not Found", f"Record with reference number {reference_number} not found")
        elif found =="Maybe":
            workbook.close()  # Close the workbook
            messagebox.showinfo("", f"Modification Cancelled for {reference_number}")

    update_button = ttk.Button(modify_window, text="Update Record", command=update_record)
    update_button.grid(row=8, column=1, columnspan=2)

def validate_login():
    # Get the username and password entered by the user
    username = username_entry.get()
    password = password_entry.get()

    # Check if the username and password are correct
    if username == "James" and password == "1234":  # Replace with your actual validation logic
        messagebox.showinfo("Login Successful", "Welcome, Admin!")
        # If login successful, open the main application window
        root.deiconify()  # Show the main window
        login_window.destroy()  # Close the login window
    else:
        messagebox.showerror("Login Failed", "Invalid username or password")

# Function to open the login page
def open_login_page():
    # Create a new window for the login page
    global login_window
    login_window = tk.Toplevel()
    login_window.title("Login Page")

    # Create labels and entry widgets for username and password
    username_label = ttk.Label(login_window, text="Username:")
    username_label.grid(row=0, column=0, padx=5, pady=5)
    global username_entry
    username_entry = ttk.Entry(login_window)
    username_entry.grid(row=0, column=1, padx=5, pady=5)

    password_label = ttk.Label(login_window, text="Password:")
    password_label.grid(row=1, column=0, padx=5, pady=5)
    global password_entry
    password_entry = ttk.Entry(login_window, show="*")  # Show '*' for password
    password_entry.grid(row=1, column=1, padx=5, pady=5)

    # Create login button
    login_button = ttk.Button(login_window, text="Login", command=validate_login)
    login_button.grid(row=2, column=0, columnspan=2, pady=10)

# Create main window
root = tk.Tk()
root.title("Index 2")
root.geometry("900x600")  # Set the size of the main window

# Hide the main window initially
root.withdraw()

# Create and pack widgets for main window
view_button = ttk.Button(root, text="View Records", command=view)
view_button.grid(row=0, column=0)

sort_button = ttk.Button(root, text="Sort Records", command=sort_workbook)
sort_button.grid(row=0, column=1)

search_button = ttk.Button(root, text="Search Record", command=search_popup)
search_button.grid(row=1, column=0)

insert_button = ttk.Button(root, text="Insert Record", command=insert_popup)
insert_button.grid(row=1, column=1)

# Create output text widget
output_text = tk.Text(root, height=10, width=100)  # Adjust the width and height as needed
output_text.grid(row=3, column=0, columnspan=2, sticky="nsew")  # Fill both horizontally and vertically

# Create a scrollbar
scrollbar = tk.Scrollbar(root, orient="vertical", command=output_text.yview)
scrollbar.grid(row=3, column=2, sticky='ns')  # Stick to the top and bottom

output_text.config(yscrollcommand=scrollbar.set)

# Configure row and column weights to make the output text widget fill the remaining space
root.rowconfigure(3, weight=1)
root.columnconfigure(0, weight=1)
root.columnconfigure(1, weight=1)

# Open the login page
open_login_page()

root.mainloop()
